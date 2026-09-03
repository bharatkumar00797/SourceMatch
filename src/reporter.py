"""
SourceMatch - Reporting Module
Generates professional Excel and text audit reports from comparison results.
"""

import os
from datetime import datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from comparator import ComparisonResult


class Reporter:
    """
    Creates clean, professional audit reports from a ComparisonResult.
    """

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.section_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        self.good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.warn_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def _style_header_row(self, ws, row: int, col_count: int):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border

    def generate_excel_report(self, result: ComparisonResult,
                              source_name: str = "Source Documents",
                              target_name: str = "Compiled File",
                              filename: str = "SourceMatch_Audit_Report.xlsx") -> str:
        """
        Generate a professional Excel audit report.
        """
        wb = Workbook()

        # ===== Sheet 1: Summary =====
        ws = wb.active
        ws.title = "Summary"

        ws["A1"] = "SourceMatch — Data Integrity Audit Report"
        ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells("A1:D1")

        ws["A2"] = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
        ws["A2"].font = Font(italic=True, size=10, color="666666")

        ws["A4"] = "Source"
        ws["B4"] = source_name
        ws["A5"] = "Target"
        ws["B5"] = target_name

        # Metrics table
        ws["A7"] = "Metric"
        ws["B7"] = "Value"
        self._style_header_row(ws, 7, 2)

        metrics = [
            ("Unique numbers in Source (originals)", result.source_total),
            ("Unique numbers in Target (compiled)", result.target_total),
            ("Numbers successfully matched", result.matched_count),
            ("Numbers missing in Target", result.missing_count),
            ("Numbers only in Target (extra)", result.extra_count),
        ]

        for i, (label, value) in enumerate(metrics, start=8):
            ws.cell(row=i, column=1, value=label).border = self.thin_border
            cell = ws.cell(row=i, column=2, value=value)
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center")

        # Accuracy highlight
        ws["A14"] = "MATCH RATE / ACCURACY"
        ws["B14"] = f"{result.match_rate:.2f}%"
        ws["A14"].font = Font(bold=True, size=12)
        ws["B14"].font = Font(bold=True, size=14, color="006600")
        ws["A14"].fill = self.good_fill
        ws["B14"].fill = self.good_fill
        ws["A14"].border = self.thin_border
        ws["B14"].border = self.thin_border

        ws["A16"] = "Formula used:"
        ws["A17"] = "Match Rate = (Matched Numbers / Total Unique Numbers in Source) × 100"
        ws["A17"].font = Font(italic=True, size=9, color="555555")

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 18

        # ===== Sheet 2: Missing Numbers =====
        ws2 = wb.create_sheet("Missing Numbers")
        ws2["A1"] = "Numbers present in Source but missing in Target"
        ws2["A1"].font = Font(bold=True, size=12)
        ws2.merge_cells("A1:B1")

        ws2["A3"] = "#"
        ws2["B3"] = "Number"
        self._style_header_row(ws2, 3, 2)

        for i, num in enumerate(result.missing_numbers, start=4):
            ws2.cell(row=i, column=1, value=i - 3).border = self.thin_border
            cell = ws2.cell(row=i, column=2, value=num)
            cell.border = self.thin_border
            cell.fill = self.warn_fill

        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 20

        if not result.missing_numbers:
            ws2["A4"] = "None — All source numbers were found in the target."

        # ===== Sheet 3: Extra Numbers =====
        ws3 = wb.create_sheet("Extra Numbers")
        ws3["A1"] = "Numbers present only in Target (not found in Source)"
        ws3["A1"].font = Font(bold=True, size=12)
        ws3.merge_cells("A1:B1")

        ws3["A3"] = "#"
        ws3["B3"] = "Number"
        self._style_header_row(ws3, 3, 2)

        for i, num in enumerate(result.extra_numbers, start=4):
            ws3.cell(row=i, column=1, value=i - 3).border = self.thin_border
            ws3.cell(row=i, column=2, value=num).border = self.thin_border

        ws3.column_dimensions["A"].width = 8
        ws3.column_dimensions["B"].width = 20

        if not result.extra_numbers:
            ws3["A4"] = "None — No extra numbers found in the target."

        # ===== Sheet 4: Matched Numbers =====
        ws4 = wb.create_sheet("Matched Numbers")
        ws4["A1"] = "Numbers successfully matched between Source and Target"
        ws4["A1"].font = Font(bold=True, size=12)
        ws4.merge_cells("A1:B1")

        ws4["A3"] = "#"
        ws4["B3"] = "Number"
        self._style_header_row(ws4, 3, 2)

        for i, num in enumerate(result.matched_numbers, start=4):
            ws4.cell(row=i, column=1, value=i - 3).border = self.thin_border
            cell = ws4.cell(row=i, column=2, value=num)
            cell.border = self.thin_border
            cell.fill = self.good_fill

        ws4.column_dimensions["A"].width = 8
        ws4.column_dimensions["B"].width = 20

        # Save
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath

    def generate_text_report(self, result: ComparisonResult,
                             source_name: str = "Source Documents",
                             target_name: str = "Compiled File",
                             filename: str = "SourceMatch_Audit_Report.txt") -> str:
        """
        Generate a clean text audit report.
        """
        lines = []
        lines.append("=" * 70)
        lines.append("SourceMatch — Data Integrity Audit Report")
        lines.append("=" * 70)
        lines.append(f"Generated on : {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
        lines.append(f"Source       : {source_name}")
        lines.append(f"Target       : {target_name}")
        lines.append("")
        lines.append("-" * 70)
        lines.append("SUMMARY")
        lines.append("-" * 70)
        lines.append(f"Unique numbers in Source (originals) : {result.source_total:,}")
        lines.append(f"Unique numbers in Target (compiled)  : {result.target_total:,}")
        lines.append(f"Numbers successfully matched         : {result.matched_count:,}")
        lines.append(f"Numbers missing in Target            : {result.missing_count:,}")
        lines.append(f"Numbers only in Target (extra)       : {result.extra_count:,}")
        lines.append("")
        lines.append(f"MATCH RATE / ACCURACY                : {result.match_rate:.2f}%")
        lines.append("")
        lines.append("Formula: (Matched Numbers / Total Unique Numbers in Source) × 100")
        lines.append("")

        lines.append("-" * 70)
        lines.append(f"MISSING NUMBERS ({result.missing_count})")
        lines.append("-" * 70)
        if result.missing_numbers:
            # Print in rows of 8
            for i in range(0, len(result.missing_numbers), 8):
                chunk = result.missing_numbers[i:i+8]
                lines.append(", ".join(chunk))
        else:
            lines.append("None — All source numbers were found in the target.")

        lines.append("")
        lines.append("-" * 70)
        lines.append(f"EXTRA NUMBERS ({result.extra_count})")
        lines.append("-" * 70)
        if result.extra_numbers:
            for i in range(0, len(result.extra_numbers), 8):
                chunk = result.extra_numbers[i:i+8]
                lines.append(", ".join(chunk))
        else:
            lines.append("None — No extra numbers found in the target.")

        lines.append("")
        lines.append("=" * 70)
        lines.append("End of Report")
        lines.append("=" * 70)

        filepath = os.path.join(self.output_dir, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        return filepath

    def generate_all(self, result: ComparisonResult,
                     source_name: str = "Source Documents",
                     target_name: str = "Compiled File") -> dict:
        """
        Generate both Excel and text reports.
        Returns paths to the generated files.
        """
        excel_path = self.generate_excel_report(result, source_name, target_name)
        text_path = self.generate_text_report(result, source_name, target_name)

        return {
            "excel": excel_path,
            "text": text_path
        }
